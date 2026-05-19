"""Geração de Excel compartilhada (openpyxl). Usada por ambas as unidades."""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_BORDA = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _header_row(ws, cols: list[tuple[str, int]], cor_fundo: str = "1F3864"):
    for col, (nome, larg) in enumerate(cols, 1):
        c = ws.cell(row=2, column=col, value=nome)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=cor_fundo)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDA
        ws.column_dimensions[get_column_letter(col)].width = larg
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"


def _titulo_row(ws, cols: list[tuple[str, int]], titulo: str):
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, color="1F3864", size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def gerar_excel_preventivas(preventivas: list, d_ini=None, d_fim=None, nome_unidade: str = "") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Preventivas Previstas"

    cols = [
        ("Data Prevista", 16), ("Dia", 10), ("Plano", 40), ("Tipo", 18), ("Oficina", 18),
        ("Equipamento", 28), ("Setor", 30), ("OS Vinculada", 14), ("Status OS", 16),
        ("Recomendado", 30), ("Cargo", 30), ("Escala", 16),
    ]

    if d_ini and d_fim:
        periodo_str = f"{d_ini.strftime('%d/%m/%Y')} a {d_fim.strftime('%d/%m/%Y')}"
    elif d_ini:
        periodo_str = f"a partir de {d_ini.strftime('%d/%m/%Y')}"
    elif d_fim:
        periodo_str = f"até {d_fim.strftime('%d/%m/%Y')}"
    else:
        periodo_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    unidade_prefix = f"{nome_unidade} — " if nome_unidade else ""
    _titulo_row(ws, cols, f"{unidade_prefix}Preventivas Previstas — {periodo_str}")
    _header_row(ws, cols)

    for i, r in enumerate(preventivas, 3):
        has_rec = bool(r.get("recomendado"))
        fill = PatternFill("solid", fgColor="FFEBEE" if not has_rec else "FFFFFF")
        for col, val in enumerate([
            r.get("data_prev", ""),
            r.get("dia_par", ""),
            r.get("plano", ""),
            r.get("tipo", ""),
            r.get("oficina", ""),
            r.get("equipamento", "—"),
            r.get("setor", "—"),
            r.get("os_vinculada", "—"),
            r.get("os_situacao", "—"),
            r.get("recomendado") or "SEM CANDIDATO",
            r.get("cargo", ""),
            r.get("escala", ""),
        ], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = _BORDA
        ws.row_dimensions[i].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def gerar_excel_recomendacoes(resultados: list, nome_unidade: str = "") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Recomendacoes"

    cols = [
        ("Nº OS", 14), ("Data", 16), ("Turno", 14), ("Situação", 14), ("Tipo", 18),
        ("Setor", 40), ("Ativo", 28), ("Resp. Atual", 30), ("Recomendado", 30),
        ("Cargo", 36), ("Score", 10), ("Status", 18),
    ]

    unidade_prefix = f"{nome_unidade} — " if nome_unidade else ""
    _titulo_row(ws, cols, f"{unidade_prefix}Recomendações — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    _header_row(ws, cols)

    cor = {
        "Bate com atual": "E8F5E9",
        "Sugestao diferente": "FFF9C4",
        "Sem resp. atual": "E3F2FD",
        "Sem candidato": "FFEBEE",
    }

    for i, r in enumerate(resultados, 3):
        fill = PatternFill("solid", fgColor=cor.get(r.get("status", ""), "FFFFFF"))
        for col, val in enumerate([
            r["numero"],
            f"{r['data_os']} {str(r['hora_os']).zfill(2)}h",
            f"{r['turno_os']} | {r['dia_par']}",
            r["situacao"],
            r["tipo"],
            r["setor"],
            r["ativo"] or "—",
            r["responsavel_atual"] or "—",
            r["recomendado"] or "SEM CANDIDATO",
            r["cargo"] or "—",
            r["score"] if r["recomendado"] else "—",
            r["status"],
        ], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = _BORDA
        ws.row_dimensions[i].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
